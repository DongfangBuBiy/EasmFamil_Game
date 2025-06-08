from openpyxl import load_workbook


# تابعی برای خواندن فایل اکسل و دسته‌بندی کلمات بر اساس حرف اول
def categorize_words_by_first_letter(file_path, sheet_name, column_index):
    # بارگذاری فایل اکسل
    workbook = load_workbook(filename=file_path)

    # انتخاب شیت مورد نظر
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"شیت '{sheet_name}' در فایل اکسل وجود ندارد.")
    sheet = workbook[sheet_name]

    # ساخت دیکشنری شامل تمام حروف الفبا (A تا Z) با لیست‌های خالی
    categorized_words = {chr(i): [] for i in range(ord('A'), ord('Z') + 1)}

    # پیمایش روی سطرهای ستون مورد نظر
    for row in sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index, values_only=True):
        word = row[0]  # گرفتن مقدار سلول
        if isinstance(word, str):  # اطمینان از اینکه مقدار یک رشته است
            first_letter = word[0].upper()  # گرفتن حرف اول و تبدیل به حرف بزرگ
            if first_letter in categorized_words:  # اطمینان از اینکه حرف در دیکشنری وجود دارد
                categorized_words[first_letter].append(word)  # اضافه کردن کلمه به لیست مربوطه

    return categorized_words


# مسیر فایل اکسل، نام شیت و شماره ستون مورد نظر
file_path = "C:/Users/AsrarIT/Desktop/Country.xlsx"  # جایگزین کنید با مسیر فایل خود
sheet_name = "Sheet1"  # نام شیت مورد نظر
column_index = 2  # شماره ستون (برای ستون b، شماره 2 است)

# فراخوانی تابع و دریافت نتیجه
result = categorize_words_by_first_letter(file_path, sheet_name, column_index)

# چاپ نتیجه
for letter, words in result.items():
    if words:  # اگر لیست کلمات خالی نباشد
        print(f"حرف {letter}: {words}")
    else:  # اگر لیست کلمات خالی باشد
        print(f"حرف {letter}: هیچ کلمه‌ای با این حرف شروع نمی‌شود.")