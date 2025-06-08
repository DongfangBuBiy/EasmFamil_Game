from openpyxl import load_workbook
import json
import os

# تابعی برای خواندن دو ستون از فایل اکسل و ذخیره به صورت لیستی از دیکشنری‌ها
def extract_key_value_pairs(file_path, sheet_name, key_col, value_col):
    # بارگذاری فایل اکسل
    workbook = load_workbook(filename=file_path)

    # انتخاب شیت مورد نظر
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"شیت '{sheet_name}' در فایل اکسل وجود ندارد.")
    sheet = workbook[sheet_name]

    # لیستی برای ذخیره جفت‌های کلید-مقدار
    pairs = []

    # پیمایش روی سطرها (از ردیف 2 به بعد)
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=max(key_col, value_col), values_only=True):
        key = row[key_col - 1]      # ستون key (براساس شماره ستون)
        value = row[value_col - 1]  # ستون value (براساس شماره ستون)

        # حذف فاصله‌ها و تبدیل به string
        key_str = str(key).strip().replace(" ", "") if key is not None else ""
        value_str = str(value).strip().replace(" ", "") if value is not None else ""

        # فقط اضافه کن اگر هر دو مقدار وجود داشته باشد
        if key_str and value_str:
            pairs.append({
                "key": key_str,
                "value": value_str
            })

    return pairs


# تابعی برای ذخیره دیکشنری‌ها در یک فایل JSON بدون فاصله
def save_dict_to_json(data, output_directory, file_name):
    # اطمینان از وجود مسیر خروجی
    if not os.path.exists(output_directory):
        os.makedirs(output_directory)

    # ساخت مسیر کامل فایل
    output_file = os.path.join(output_directory, f"{file_name}.json")

    # ذخیره داده‌ها در فایل JSON بدون فاصله و خطوط اضافی
    with open(output_file, "w", encoding="utf-8") as json_file:
        json.dump({"data": data}, json_file, ensure_ascii=False, separators=(',', ':'))

    print(f"داده‌ها با موفقیت در فایل '{output_file}' ذخیره شدند!")


# مسیر فایل اکسل، نام شیت و شماره ستون‌ها
file_path = "C:/Users/AsrarIT/Desktop/use exsel/job.xlsx"
sheet_name = "Sheet1"

# ستون A -> value
# ستون B -> key
value_col = 1  # ستون اول (A)
key_col = 2    # ستون دوم (B)

# مسیر خروجی برای فایل JSON
output_directory = "C:/Users/AsrarIT/Desktop/flowerjason"

# نام فایل JSON (بدون .json)
file_name = input("لطفاً نام فایل JSON را وارد کنید: ").strip()

# خواندن دو ستون و تبدیل به لیستی از دیکشنری‌ها ✅
result = extract_key_value_pairs(file_path, sheet_name, key_col, value_col)  # ⬅️ اسم درست تابع

# ذخیره در فایل JSON
save_dict_to_json(result, output_directory, file_name)